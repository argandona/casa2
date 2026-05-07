# =============================================================================
# IMPORTS
# =============================================================================

import json
import traceback
import colorsys
from collections import defaultdict
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO

import pandas as pd

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.core.management.base import BaseCommand
from django.core.paginator import Paginator
from django.db import models as django_models, transaction
from django.db.models import Count, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST, require_http_methods

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .models import (
    SST, Suministro, Distrito,
    EstadoSST, EstadoSuministro, EstadoLiquidacion
)


# =============================================================================
# DASHBOARD
# =============================================================================

@login_required
def dashboard(request):
    """Dashboard principal con KPIs de SST y suministros."""

    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # --- KPIs SST ---
    sst_totales = SST.objects.count()
    sst_ejecutadas = SST.objects.filter(estado_sst__estado='EJECUTADO').count()
    sst_en_ejecucion = SST.objects.filter(estado_sst__estado='EN EJECUCIÓN').count()
    sst_admisibles = SST.objects.filter(estado_sst__estado='ADMISIBLE').count()
    sst_otros = sst_totales - (sst_ejecutadas + sst_en_ejecucion + sst_admisibles)

    # Porcentajes SST
    if sst_totales > 0:
        porcentaje_ejecutadas = round((sst_ejecutadas / sst_totales) * 100, 1)
        porcentaje_en_ejecucion = round((sst_en_ejecucion / sst_totales) * 100, 1)
        porcentaje_admisibles = round((sst_admisibles / sst_totales) * 100, 1)
        porcentaje_otros = round((sst_otros / sst_totales) * 100, 1)
    else:
        porcentaje_ejecutadas = porcentaje_en_ejecucion = porcentaje_admisibles = porcentaje_otros = 0

    # --- KPIs Suministros ---
    total_suministros = Suministro.objects.count()
    suministros_ejecutados = Suministro.objects.filter(
        estado_suministro__estado_suministro='EJECUTADO'
    ).count()
    suministros_asignados = Suministro.objects.filter(
        estado_suministro__estado_suministro='ASIGNADO'
    ).count()
    suministros_devueltos = Suministro.objects.filter(
        estado_suministro__estado_suministro='DEVUELTO'
    ).count()

    # --- Montos ---
    monto_total_proyectado = SST.objects.aggregate(
        total=Sum('monto_proyectado')
    )['total'] or Decimal('0.00')
    monto_total_real = SST.objects.aggregate(
        total=Sum('monto_real')
    )['total'] or Decimal('0.00')

    if monto_total_proyectado > 0:
        eficiencia_financiera = round((monto_total_real / monto_total_proyectado) * 100, 1)
    else:
        eficiencia_financiera = 0

    # --- Este mes ---
    sst_este_mes = SST.objects.filter(created_at__gte=inicio_mes).count()
    sst_ejecutadas_mes = SST.objects.filter(
        created_at__gte=inicio_mes,
        estado_sst__estado='EJECUTADO'
    ).count()
    monto_mes = SST.objects.filter(created_at__gte=inicio_mes).aggregate(
        total=Sum('monto_proyectado')
    )['total'] or Decimal('0.00')
    suministros_mes = Suministro.objects.filter(created_at__gte=inicio_mes).count()

    # --- Listados ---
    ultimas_sst = SST.objects.select_related(
        'distrito', 'estado_sst'
    ).order_by('-created_at')[:10]

    sst_en_ejecucion_list = SST.objects.filter(
        estado_sst__estado='EN EJECUCIÓN'
    ).select_related('distrito').order_by('-created_at')[:10]

    for sst in sst_en_ejecucion_list:
        sst.dias_transcurridos = (timezone.now().date() - sst.created_at.date()).days

    context = {
        # SST
        'sst_totales': sst_totales,
        'sst_ejecutadas': sst_ejecutadas,
        'sst_en_ejecucion': sst_en_ejecucion,
        'sst_admisibles': sst_admisibles,
        'sst_otros': sst_otros,
        # Porcentajes SST
        'porcentaje_ejecutadas': porcentaje_ejecutadas,
        'porcentaje_en_ejecucion': porcentaje_en_ejecucion,
        'porcentaje_admisibles': porcentaje_admisibles,
        'porcentaje_otros': porcentaje_otros,
        # Suministros
        'total_suministros': total_suministros,
        'suministros_ejecutados': suministros_ejecutados,
        'suministros_asignados': suministros_asignados,
        'suministros_devueltos': suministros_devueltos,
        # Financiero
        'monto_total_proyectado': monto_total_proyectado,
        'monto_total_real': monto_total_real,
        'eficiencia_financiera': eficiencia_financiera,
        # Este mes
        'sst_este_mes': sst_este_mes,
        'sst_ejecutadas_mes': sst_ejecutadas_mes,
        'monto_mes': monto_mes,
        'suministros_mes': suministros_mes,
        # Listados
        'ultimas_sst': ultimas_sst,
        'sst_en_ejecucion_list': sst_en_ejecucion_list,
    }

    return render(request, 'gestion/dashboard.html', context)


# =============================================================================
# SST
# =============================================================================

@login_required
def sst_list(request):
    ssts = SST.objects.select_related('distrito', 'estado_sst').all()
    distritos = Distrito.objects.all()

    context = {
        'ssts': ssts,
        'distritos': distritos,
    }
    return render(request, 'gestion/sst.html', context)


@login_required
def buscar_sst(request):
    """Autocompletado de búsqueda de SST."""
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'success': True, 'data': []})

    try:
        sst_qs = SST.objects.filter(
            Q(sst__icontains=query) | Q(direccion__icontains=query)
        ).select_related('distrito').order_by('sst')[:10]

        results = []
        for sst in sst_qs:
            total_original = sst.suministros.filter(tipo_suministro='ORIGINAL').count()
            total_adicional = sst.suministros.filter(tipo_suministro='ADICIONAL').count()
            results.append({
                'id': sst.id,
                'sst': sst.sst,
                'direccion': sst.direccion,
                'distrito': sst.distrito.nombre_distrito if sst.distrito else '',
                'estado': sst.estado_sst.estado if sst.estado_sst else '',
                'total_suministros': f"{total_original}/{total_adicional}",
            })

        return JsonResponse({'success': True, 'data': results})

    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Error en buscar_sst: {e}")
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor',
            'data': [],
        }, status=500)


@login_required
def obtener_info_sst(request, sst_id):
    """Información de una SST para usar en modales."""
    try:
        sst = SST.objects.select_related('distrito', 'estado_sst').get(id=sst_id)

        total_suministros = sst.suministros.count()
        suministros_adicionales = sst.suministros.filter(tipo_suministro='ADICIONAL').count()

        # Sugerencia de número para el siguiente suministro adicional
        ultimo_adicional = sst.suministros.filter(
            tipo_suministro='ADICIONAL'
        ).order_by('-suministro').first()

        try:
            if ultimo_adicional:
                ultimo_num = ultimo_adicional.suministro.split('-')[-1]
                if ultimo_num.startswith('AD'):
                    num = int(ultimo_num[2:]) + 1
                    sugerencia = f"{sst.sst}-AD{num:03d}"
                else:
                    sugerencia = f"{sst.sst}-AD001"
            else:
                sugerencia = f"{sst.sst}-AD001"
        except Exception:
            sugerencia = f"{sst.sst}-AD001"

        data = {
            'sst': sst.sst,
            'direccion': sst.direccion,
            'distrito': sst.distrito.nombre_distrito if sst.distrito else 'Sin distrito',
            'estado': sst.estado_sst.estado if sst.estado_sst else 'Sin estado',
            'total_suministros': total_suministros,
            'suministros_adicionales': suministros_adicionales,
            'sugerencia_suministro': sugerencia,
        }
        return JsonResponse({'success': True, 'data': data})

    except SST.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'SST no encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# =============================================================================
# SUMINISTROS — listado
# =============================================================================

@login_required
def suministro_list(request):
    # --- Filtros ---
    sst_filter = request.GET.get('sst', '')
    distrito_filter = request.GET.get('distrito', '')
    estado_filter = request.GET.get('estado', '')
    search = request.GET.get('search', '')
    obs_lds_filter = request.GET.get('obs_lds', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    todas_sst = SST.objects.select_related('distrito', 'estado_sst').order_by('sst')

    suministros = Suministro.objects.select_related(
        'sst', 'distrito', 'estado_suministro'
    ).all()

    if sst_filter:
        suministros = suministros.filter(sst__sst__icontains=sst_filter)
    if distrito_filter:
        suministros = suministros.filter(distrito__nombre_distrito=distrito_filter)
    if estado_filter:
        suministros = suministros.filter(
            estado_suministro__estado_suministro=estado_filter
        )
    if search:
        suministros = suministros.filter(
            Q(suministro__icontains=search) |
            Q(medidor__icontains=search) |
            Q(direccion__icontains=search)
        )
    if obs_lds_filter:
        suministros = suministros.filter(Observacion_LDS__icontains=obs_lds_filter)
        
        
   

# ← AGREGAR ESTO:
    if fecha_desde:
        try:
            suministros = suministros.filter(
                fecha_ejecucion__gte=datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        )
        except ValueError:
            pass
    if fecha_hasta:
        try:
            suministros = suministros.filter(
            fecha_ejecucion__lte=datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        )
        except ValueError:
            pass        

    # --- Paginación ---
    paginator = Paginator(suministros, 50)
    page_obj = paginator.get_page(request.GET.get('page'))

    # --- Catálogos (con cache para distritos) ---
    distritos = cache.get('distritos')
    if not distritos:
        distritos = Distrito.objects.all()
        cache.set('distritos', distritos, 3600)

    estados = EstadoSuministro.objects.all()
    ssts = SST.objects.all()

    # --- Estadísticas SST ---
    sst_por_estado = {
        i['estado_sst__estado']: i['cantidad']
        for i in SST.objects.values('estado_sst__estado').annotate(cantidad=Count('id'))
    }

    # --- Estadísticas Suministros ---
    suministros_por_estado = {
        i['estado_suministro__estado_suministro']: i['cantidad']
        for i in Suministro.objects.values(
            'estado_suministro__estado_suministro'
        ).annotate(cantidad=Count('id'))
    }

    context = {
        'suministros': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'distritos': distritos,
        'estados': estados,
        'ssts': ssts,
        'todas_sst': todas_sst,
        # Filtros activos
        'sst_filter': sst_filter,
        'distrito_filter': distrito_filter,
        'estado_filter': estado_filter,
        'search': search,
        'obs_lds_filter': obs_lds_filter,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        # Stats SST
        'sst_ejecutado': sst_por_estado.get('Ejecutado', 0),
        'sst_admisible': sst_por_estado.get('Admisible', 0),
        'sst_en_ejecucion': sst_por_estado.get('En Ejecución', 0),
        'sst_pendiente': sst_por_estado.get('Pendiente', 0),
        # Stats Suministros
        'sum_asignado': (suministros_por_estado.get('ASIGNADO', 0)
                         + suministros_por_estado.get('Asignado', 0)),
        'sum_devuelto': (suministros_por_estado.get('DEVUELTO', 0)
                         + suministros_por_estado.get('Devuelto', 0)),
        'sum_ejecutado': (suministros_por_estado.get('EJECUTADO', 0)
                          + suministros_por_estado.get('Ejecutado', 0)),
        'total_suministros': suministros.count(),
    }

    return render(request, 'gestion/suministros.html', context)


# =============================================================================
# SUMINISTROS — CRUD
# =============================================================================

@login_required
@require_POST
def actualizar_suministro(request, suministro_id):
    try:
        suministro = Suministro.objects.get(id=suministro_id)
        data = json.loads(request.body)
        monto_anterior = suministro.monto

        # Estado
        estado_id = data.get('estado_suministro')
        if estado_id:
            try:
                suministro.estado_suministro = EstadoSuministro.objects.get(id=estado_id)
            except EstadoSuministro.DoesNotExist:
                pass

        # Fecha de ejecución
        fecha_ejecucion = data.get('fecha_ejecucion')
        if fecha_ejecucion:
            suministro.fecha_ejecucion = datetime.strptime(fecha_ejecucion, '%Y-%m-%d').date()

        # Ejecutado por
        ejecutado_por = data.get('ejecutado_por')
        if ejecutado_por:
            suministro.ejecutado_por = ejecutado_por

        # Monto
        monto = data.get('monto')
        if monto is not None:
            try:
                suministro.monto = Decimal(str(monto))
            except (ValueError, TypeError):
                pass

        # Observación
        observacion = data.get('observacion_contratista')
        if observacion is not None:
            suministro.observacion_contratista = observacion

        suministro.save()

        if monto_anterior != suministro.monto:
            suministro.sst.actualizar_monto_total()

        suministro.sst.actualizar_estado_segun_suministros()

        return JsonResponse({'success': True, 'message': 'Suministro actualizado correctamente'})

    except Suministro.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Suministro no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@require_POST
def eliminar_suministro(request, suministro_id):
    try:
        suministro = Suministro.objects.get(id=suministro_id)
        sst = suministro.sst
        suministro.delete()
        sst.actualizar_estado_segun_suministros()

        return JsonResponse({
            'success': True,
            'message': f'Suministro {suministro.suministro} eliminado correctamente',
        })

    except Suministro.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Suministro no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=400)


@login_required
@csrf_exempt
def agregar_suministro_adicional(request):
    """Agrega un suministro adicional vía AJAX."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)

    try:
        data = json.loads(request.body)

        sst_id = data.get('sst_id')
        if not sst_id:
            return JsonResponse({'success': False, 'message': 'Debe seleccionar una SST'}, status=400)

        motivo_adicional = data.get('motivo_adicional', '').strip()
        if not motivo_adicional:
            return JsonResponse({
                'success': False,
                'message': 'El motivo del suministro adicional es obligatorio',
            }, status=400)

        try:
            sst = SST.objects.get(id=sst_id)
        except SST.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'SST no encontrada'}, status=404)

        # Número de suministro
        suministro_num = data.get('suministro', '').strip()
        if not suministro_num:
            count_adicionales = Suministro.objects.filter(
                sst=sst, tipo_suministro='ADICIONAL'
            ).count()
            suministro_num = f"{sst.sst}-AD{count_adicionales + 1:03d}"

        # Item
        ultimo_suministro = Suministro.objects.filter(sst=sst).order_by('-item').first()
        nuevo_item = (ultimo_suministro.item + 1) if ultimo_suministro else 1

        # Estado
        estado_sum = None
        estado_id = data.get('estado_suministro')
        if estado_id:
            try:
                estado_sum = EstadoSuministro.objects.get(id=estado_id)
            except EstadoSuministro.DoesNotExist:
                pass

        if not estado_sum:
            estado_sum, _ = EstadoSuministro.objects.get_or_create(
                estado_suministro='PENDIENTE',
                defaults={'descripcion': 'Pendiente de ejecución', 'color': '#FFA500'},
            )

        # Monto
        monto = Decimal('0.00')
        monto_data = data.get('monto')
        if monto_data:
            try:
                monto = Decimal(str(monto_data))
            except Exception:
                pass

        suministro = Suministro(
            sst=sst,
            tipo_suministro='ADICIONAL',
            item=nuevo_item,
            no_ot=sst.sst,
            suministro=suministro_num,
            monto=monto,
            estado_suministro=estado_sum,
            direccion=sst.direccion,
            distrito=sst.distrito,
            motivo_adicional=motivo_adicional,
            fecha_ejecucion=data.get('fecha_ejecucion') or None,
            ejecutado_por=data.get('ejecutado_por', '').strip() or None,
            contacto=data.get('contacto', '').strip() or None,
            telefono=data.get('telefono', '').strip() or None,
            solicitado_por=data.get('solicitado_por', '').strip() or None,
            observacion_contratista=data.get('observacion_contratista', '').strip() or None,
            fecha_identificacion=datetime.now().date(),
        )
        suministro.save()
        sst.actualizar_estado_segun_suministros()

        return JsonResponse({
            'success': True,
            'message': f'Suministro adicional "{suministro_num}" creado exitosamente',
            'suministro_id': suministro.id,
            'suministro_numero': suministro_num,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


# =============================================================================
# SUMINISTROS — importar / exportar Excel
# =============================================================================

@login_required
def cargar_excel_suministros(request):
    """Carga masiva de suministros desde un archivo Excel."""
    if not (request.method == 'POST' and request.FILES.get('excel_file')):
        return redirect('sst_list')

    try:
        df = pd.read_excel(request.FILES['excel_file'])
        df.columns = df.columns.str.strip()

        columnas_criticas = ['Item', 'No. OT', 'Suministro', 'Dirección', 'Distrito']
        columnas_faltantes = [c for c in columnas_criticas if c not in df.columns]
        if columnas_faltantes:
            messages.error(request, f'❌ Faltan columnas: {", ".join(columnas_faltantes)}')
            return redirect('sst_list')

        estado_sst_default, _ = EstadoSST.objects.get_or_create(
            estado='PENDIENTE',
            defaults={'descripcion': 'Estado por defecto', 'color': '#FFA500'},
        )
        estado_suministro_default, _ = EstadoSuministro.objects.get_or_create(
            estado_suministro='PENDIENTE',
            defaults={'descripcion': 'Estado por defecto', 'color': '#FFA500'},
        )

        ssts_creadas = suministros_creados = suministros_actualizados = 0
        errores = []

        with transaction.atomic():
            ssts_procesadas = set()

            for ot in df['No. OT'].dropna().unique():
                try:
                    suministros_ot = df[df['No. OT'] == ot]
                    if len(suministros_ot) == 0:
                        continue

                    codigo_sst = str(ot).strip()[:7] if len(str(ot).strip()) >= 7 else str(ot).strip()

                    distrito_nombre = (
                        suministros_ot['Distrito'].mode()[0]
                        if not suministros_ot['Distrito'].mode().empty
                        else suministros_ot['Distrito'].iloc[0]
                    )
                    distrito_principal, _ = Distrito.objects.get_or_create(
                        nombre_distrito=str(distrito_nombre).strip()
                    )

                    direccion_principal = (
                        str(suministros_ot['Dirección'].iloc[0]).strip()
                        if pd.notna(suministros_ot['Dirección'].iloc[0])
                        else 'Sin dirección'
                    )

                    sst, created = SST.objects.get_or_create(
                        sst=codigo_sst,
                        defaults={
                            'direccion': direccion_principal,
                            'distrito': distrito_principal,
                            'estado_sst': estado_sst_default,
                            'monto_proyectado': Decimal('0.00'),
                        },
                    )
                    if created:
                        ssts_creadas += 1

                    ssts_procesadas.add(sst.id)

                    for index, row in suministros_ot.iterrows():
                        try:
                            item = int(row['Item']) if pd.notna(row['Item']) else 0
                            no_ot = str(row['No. OT']).strip() if pd.notna(row['No. OT']) else ''
                            suministro_codigo = str(row['Suministro']).strip() if pd.notna(row['Suministro']) else ''
                            medidor = str(row['Medidor']).strip() if pd.notna(row['Medidor']) else None
                            marca_medidor = str(row['Marca']).strip() if pd.notna(row['Marca']) else None
                            fase = str(row['Fase']).strip() if pd.notna(row['Fase']) else None
                            potencia = str(row['Potencia']).strip() if pd.notna(row['Potencia']) else None
                            direccion = str(row['Dirección']).strip() if pd.notna(row['Dirección']) else ''
                            distrito_nombre = str(row['Distrito']).strip() if pd.notna(row['Distrito']) else ''
                            observacion_lds = str(row.get('Observación LDS', '')).strip() if pd.notna(row.get('Observación LDS')) else None
                            latitud = row.get('Latitud')
                            longitud = row.get('Longitud')
                            contacto = str(row.get('Contacto', '')).strip() if pd.notna(row.get('Contacto')) else None
                            telefono = str(row.get('Teléfono', '')).strip() if pd.notna(row.get('Teléfono')) else None
                            ejecutado_por = str(row.get('Ejecutado Por', '')).strip() if pd.notna(row.get('Ejecutado Por')) else None
                            observacion_contratista = str(row.get('Observación Contratista', '')).strip() if pd.notna(row.get('Observación Contratista')) else None

                            fecha_primer_envio = pd.to_datetime(row.get('Fecha Primer Envío'), errors='coerce') if pd.notna(row.get('Fecha Primer Envío')) else None
                            fecha_programada = pd.to_datetime(row.get('Fec. Prog.'), errors='coerce') if pd.notna(row.get('Fec. Prog.')) else None
                            fecha_ejecucion = pd.to_datetime(row.get('Fecha Ejecución'), errors='coerce') if pd.notna(row.get('Fecha Ejecución')) else None

                            hora_inicio_prog = hora_fin_prog = None
                            if pd.notna(row.get('Hor. Prog.')):
                                try:
                                    hora_inicio_prog = pd.to_datetime(str(row['Hor. Prog.']), format='%H:%M:%S', errors='coerce').time()
                                except Exception:
                                    pass
                            if pd.notna(row.get('Hor. Fin. Prog.')):
                                try:
                                    hora_fin_prog = pd.to_datetime(str(row['Hor. Fin. Prog.']), format='%H:%M:%S', errors='coerce').time()
                                except Exception:
                                    pass

                            distrito_suministro, _ = Distrito.objects.get_or_create(nombre_distrito=distrito_nombre)

                            estado_nombre = str(row.get('Estado', '')).strip() if pd.notna(row.get('Estado')) else ''
                            if estado_nombre:
                                estado_sum, _ = EstadoSuministro.objects.get_or_create(
                                    estado_suministro=estado_nombre,
                                    defaults={'descripcion': '', 'color': '#007bff'},
                                )
                            else:
                                estado_sum = estado_suministro_default

                            campos = dict(
                                item=item, no_ot=no_ot, medidor=medidor,
                                marca_medidor=marca_medidor, fase=fase, potencia=potencia,
                                direccion=direccion, distrito=distrito_suministro,
                                latitud=latitud, longitud=longitud, contacto=contacto,
                                telefono=telefono, fecha_primer_envio=fecha_primer_envio,
                                fecha_programada=fecha_programada,
                                hora_inicio_programada=hora_inicio_prog,
                                hora_fin_programada=hora_fin_prog,
                                fecha_ejecucion=fecha_ejecucion,
                                ejecutado_por=ejecutado_por, estado_suministro=estado_sum,
                                observacion_contratista=observacion_contratista,
                                Observacion_LDS=observacion_lds,
                            )

                            existente = Suministro.objects.filter(
                                sst=sst, suministro=suministro_codigo
                            ).first()

                            if existente:
                                for attr, value in campos.items():
                                    setattr(existente, attr, value)
                                existente.save()
                                suministros_actualizados += 1
                            else:
                                Suministro.objects.create(
                                    sst=sst, suministro=suministro_codigo, **campos
                                )
                                suministros_creados += 1

                        except Exception as e:
                            errores.append(f"Fila {index + 2}: {e}")

                except Exception as e:
                    errores.append(f"OT {ot}: {e}")

            for sst_id in ssts_procesadas:
                try:
                    SST.objects.get(id=sst_id).actualizar_estado_segun_suministros()
                except Exception as e:
                    errores.append(f"Error actualizando SST {sst_id}: {e}")

        partes = []
        if ssts_creadas:
            partes.append(f'{ssts_creadas} SST creadas')
        if suministros_creados:
            partes.append(f'{suministros_creados} suministros creados')
        if suministros_actualizados:
            partes.append(f'{suministros_actualizados} suministros actualizados')
        if partes:
            messages.success(request, f'✅ {", ".join(partes)}.')
        if errores:
            messages.warning(request, f'⚠️ {len(errores)} errores encontrados.')
            for error in errores[:5]:
                messages.warning(request, error)

    except Exception as e:
        messages.error(request, f'❌ Error: {e}')

    return redirect('sst_list')


@login_required
@require_http_methods(["POST"])
def importar_excel_suministros(request):
    """Actualización masiva de suministros desde Excel (SST + Suministro + Estado obligatorios)."""
    if 'archivo' not in request.FILES:
        return JsonResponse({'success': False, 'message': 'No se ha seleccionado ningún archivo'})

    archivo = request.FILES['archivo']
    if not archivo.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': 'El archivo debe ser un Excel (.xlsx o .xls)'})

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        headers_normalized = {h.strip().lower(): idx for idx, h in enumerate(headers) if h}

        required_columns = ['SST', 'Suministro', 'Estado']
        missing_columns = [c for c in required_columns if c.lower() not in headers_normalized]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Faltan columnas requeridas: {", ".join(missing_columns)}',
            })

        col_sst = headers_normalized.get('sst')
        col_suministro = headers_normalized.get('suministro')
        col_estado = headers_normalized.get('estado')
        col_monto = headers_normalized.get('monto')
        col_ejecutado_por = (headers_normalized.get('ejecutado por')
                             or headers_normalized.get('ejecutado_por'))
        col_fecha = (headers_normalized.get('fecha de ejecucion')
                     or headers_normalized.get('fecha_ejecucion')
                     or headers_normalized.get('fecha de ejecución'))

        actualizados = filas_procesadas = 0
        errores = []
        ssts_afectadas = set()

        with transaction.atomic():
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                filas_procesadas += 1

                try:
                    sst_code = str(row[col_sst]).strip() if row[col_sst] else None
                    suministro_code = str(row[col_suministro]).strip() if row[col_suministro] else None
                    estado_nombre = str(row[col_estado]).strip() if row[col_estado] else None

                    if not (sst_code and suministro_code and estado_nombre):
                        errores.append({'fila': row_num, 'error': 'SST, Suministro y Estado son obligatorios'})
                        continue

                    try:
                        sst = SST.objects.get(sst=sst_code)
                        ssts_afectadas.add(sst.id)
                    except SST.DoesNotExist:
                        errores.append({'fila': row_num, 'sst': sst_code,
                                        'error': f'SST "{sst_code}" no existe en el sistema'})
                        continue

                    try:
                        suministro = Suministro.objects.select_related(
                            'sst', 'estado_suministro'
                        ).get(sst=sst, suministro=suministro_code)
                    except Suministro.DoesNotExist:
                        errores.append({'fila': row_num, 'sst': sst_code, 'suministro': suministro_code,
                                        'error': f'Suministro "{suministro_code}" no existe en SST "{sst_code}"'})
                        continue

                    try:
                        estado = EstadoSuministro.objects.get(
                            estado_suministro__iexact=estado_nombre
                        )
                    except EstadoSuministro.DoesNotExist:
                        validos = ', '.join(
                            EstadoSuministro.objects.values_list('estado_suministro', flat=True)
                        )
                        errores.append({'fila': row_num, 'sst': sst_code, 'suministro': suministro_code,
                                        'error': f'Estado "{estado_nombre}" no existe. Válidos: {validos}'})
                        continue

                    suministro.estado_suministro = estado
                    campos_actualizar = ['estado_suministro']

                    if col_monto is not None and row[col_monto]:
                        try:
                            monto = Decimal(str(row[col_monto]).strip().replace(',', '.'))
                            if monto < 0:
                                errores.append({'fila': row_num, 'sst': sst_code,
                                                'suministro': suministro_code,
                                                'error': 'El monto no puede ser negativo',
                                                'advertencia': True})
                            else:
                                suministro.monto = monto
                                campos_actualizar.append('monto')
                        except (InvalidOperation, ValueError):
                            errores.append({'fila': row_num, 'sst': sst_code,
                                            'suministro': suministro_code,
                                            'error': f'Monto inválido: "{row[col_monto]}"',
                                            'advertencia': True})

                    if col_ejecutado_por is not None and row[col_ejecutado_por]:
                        suministro.ejecutado_por = str(row[col_ejecutado_por]).strip()
                        campos_actualizar.append('ejecutado_por')

                    if col_fecha is not None and row[col_fecha]:
                        fecha_value = row[col_fecha]
                        try:
                            if isinstance(fecha_value, datetime):
                                suministro.fecha_ejecucion = fecha_value.date()
                                campos_actualizar.append('fecha_ejecucion')
                            elif isinstance(fecha_value, str):
                                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                    try:
                                        suministro.fecha_ejecucion = datetime.strptime(
                                            fecha_value.strip(), fmt
                                        ).date()
                                        campos_actualizar.append('fecha_ejecucion')
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    errores.append({'fila': row_num, 'sst': sst_code,
                                                    'suministro': suministro_code,
                                                    'error': f'Fecha inválida: "{fecha_value}". Use DD/MM/YYYY',
                                                    'advertencia': True})
                        except Exception as e:
                            errores.append({'fila': row_num, 'sst': sst_code,
                                            'suministro': suministro_code,
                                            'error': f'Error procesando fecha: {e}',
                                            'advertencia': True})

                    suministro.save(update_fields=campos_actualizar)
                    actualizados += 1

                except Exception as e:
                    errores.append({'fila': row_num, 'error': f'Error inesperado: {e}'})

            for sst_id in ssts_afectadas:
                try:
                    s = SST.objects.get(id=sst_id)
                    s.actualizar_monto_total()
                    s.actualizar_estado_segun_suministros()
                except Exception as e:
                    print(f"⚠️ Error actualizando SST {sst_id}: {e}")

        mensaje = (f'Proceso completado: {actualizados} suministros actualizados '
                   f'de {filas_procesadas} filas procesadas')
        if errores:
            mensaje += f'. Se encontraron {len(errores)} errores o advertencias.'

        messages.success(request, mensaje)
        return redirect('suministro_list')

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error procesando el archivo: {e}',
            'traceback': traceback.format_exc(),
        })


@login_required
def descargar_excel_suministros(request):
    """Exporta los suministros filtrados a Excel."""
    try:
        sst_filter = request.GET.get('sst', '')
        distrito_filter = request.GET.get('distrito', '')
        estado_filter = request.GET.get('estado', '')
        search = request.GET.get('search', '')
        obs_lds_filter = request.GET.get('obs_lds', '')
        fecha_desde = request.GET.get('fecha_desde', '')
        fecha_hasta = request.GET.get('fecha_hasta', '')

        suministros = Suministro.objects.select_related(
            'sst', 'distrito', 'estado_suministro'
        ).all()

        if sst_filter:
            suministros = suministros.filter(sst__sst__icontains=sst_filter)
        if distrito_filter:
            suministros = suministros.filter(distrito__nombre_distrito=distrito_filter)
        if estado_filter:
            suministros = suministros.filter(estado_suministro__estado_suministro=estado_filter)
        if search:
            suministros = suministros.filter(
                Q(suministro__icontains=search) |
                Q(medidor__icontains=search) |
                Q(direccion__icontains=search)
            )
        if obs_lds_filter:
            suministros = suministros.filter(Observacion_LDS__icontains=obs_lds_filter)   
            
            
            
        if fecha_desde:
            try:
                fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                suministros = suministros.filter(fecha_ejecucion__gte=fecha_desde_obj)
            except ValueError:
                pass
        
        if fecha_hasta:
            try:
                fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                suministros = suministros.filter(fecha_ejecucion__lte=fecha_hasta_obj)
            except ValueError:
                pass     

        data = []
        for s in suministros:
            data.append({
                'ITEM': s.item,
                'SST': s.sst.sst if s.sst else '',
                'SUMINISTRO': s.suministro,
                'TIPO': s.tipo_suministro,
                'MEDIDOR': s.medidor or '',
                'MARCA_MEDIDOR': s.marca_medidor or '',
                'FASE': s.fase or '',
                'POTENCIA': s.potencia or '',
                'ESTADO': s.estado_suministro.estado_suministro if s.estado_suministro else '',
                'DIRECCIÓN': s.direccion,
                'DISTRITO': s.distrito.nombre_distrito if s.distrito else '',
                'CONTACTO': s.contacto or '',
                'TELÉFONO': s.telefono or '',
                'LATITUD': s.latitud or '',
                'LONGITUD': s.longitud or '',
                'FECHA_PRIMER_ENVIO': s.fecha_primer_envio.strftime('%d/%m/%Y') if s.fecha_primer_envio else '',
                'FECHA_PROGRAMADA': s.fecha_programada.strftime('%d/%m/%Y') if s.fecha_programada else '',
                'HORA_INICIO_PROG': s.hora_inicio_programada.strftime('%H:%M') if s.hora_inicio_programada else '',
                'HORA_FIN_PROG': s.hora_fin_programada.strftime('%H:%M') if s.hora_fin_programada else '',
                'FECHA_EJECUCIÓN': s.fecha_ejecucion.strftime('%d/%m/%Y') if s.fecha_ejecucion else '',
                'EJECUTADO_POR': s.ejecutado_por or '',
                'OBSERVACIÓN_CONTRATISTA': s.observacion_contratista or '',
                'OBSERVACIÓN_LDS': s.Observacion_LDS or '',
                'MOTIVO_ADICIONAL': s.motivo_adicional or '',
                'FECHA_IDENTIFICACIÓN': s.fecha_identificacion.strftime('%d/%m/%Y') if s.fecha_identificacion else '',
                'SOLICITADO_POR': s.solicitado_por or '',
            })

        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Suministros', index=False)
            ws = writer.sheets['Suministros']
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output.seek(0)
        filename = f"suministros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        messages.error(request, f'Error al generar Excel: {e}')
        return redirect('suministro_list')


@login_required
def descargar_plantilla_importacion(request):
    """Genera una plantilla Excel de ejemplo para importar suministros."""
    from datetime import date as date_type

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['SST', 'Suministro', 'Estado', 'Monto', 'Ejecutado por', 'Fecha de ejecucion']
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ejemplos = [
        ['SST-001', 'SST-001-001', 'Ejecutado', '150.00', 'Juan Pérez', '15/01/2026'],
        ['SST-001', 'SST-001-002', 'Ejecutado', '200.50', 'María García', '16/01/2026'],
        ['SST-002', 'SST-002-001', 'Pendiente', '', '', ''],
        ['SST-002', 'SST-002-AD001', 'Asignado', '180.00', 'Carlos López', '17/01/2026'],
    ]

    for row_num, row_data in enumerate(ejemplos, 2):
        ws.append(row_data)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = example_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    for col_num, width in enumerate([15, 20, 15, 12, 20, 18], 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Hoja de instrucciones
    ws_i = wb.create_sheet("Instrucciones")
    instrucciones = [
        ("INSTRUCCIONES PARA IMPORTAR SUMINISTROS", ""),
        ("", ""),
        ("Columnas Obligatorias:", ""),
        ("✓ SST", "Código de la SST (debe existir en el sistema)"),
        ("✓ Suministro", "Número de suministro (debe existir en el sistema)"),
        ("✓ Estado", "Estado del suministro (Ejecutado, Pendiente, Asignado, etc.)"),
        ("", ""),
        ("Columnas Opcionales:", ""),
        ("○ Monto", "Monto en soles (número con hasta 2 decimales)"),
        ("○ Ejecutado por", "Nombre del técnico que ejecutó el trabajo"),
        ("○ Fecha de ejecucion", "Formato: DD/MM/YYYY"),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("1.", "Los suministros DEBEN existir previamente en el sistema"),
        ("2.", "Solo se actualizarán los campos incluidos en el Excel"),
        ("3.", "Los campos vacíos NO modificarán los datos existentes"),
        ("4.", "El monto debe ser un número positivo"),
        ("5.", "La fecha debe estar en formato DD/MM/YYYY"),
        ("6.", "El estado debe coincidir exactamente con los estados del sistema"),
        ("", ""),
        ("Estados válidos:", "Consultar en el sistema los estados disponibles"),
    ]

    titulos_negrita = {
        "Columnas Obligatorias:", "Columnas Opcionales:",
        "NOTAS IMPORTANTES:", "Estados válidos:",
    }

    for row_num, (col1, col2) in enumerate(instrucciones, 1):
        ws_i.cell(row=row_num, column=1, value=col1)
        ws_i.cell(row=row_num, column=2, value=col2)
        if row_num == 1:
            ws_i.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="4472C4")
        elif col1 in titulos_negrita:
            ws_i.cell(row=row_num, column=1).font = Font(bold=True, size=11)

    ws_i.column_dimensions['A'].width = 25
    ws_i.column_dimensions['B'].width = 60

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=plantilla_importacion_suministros_{date_type.today()}.xlsx'
    )
    wb.save(response)
    return response


@login_required
def descargar_excel_modulo(request):
    pass  # TODO: implementar


# =============================================================================
# REPORTES
# =============================================================================

@login_required
def reporte_productividad(request):
    """Reporte en formato MATRIZ: fechas en filas, ejecutores en columnas."""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    suministros = Suministro.objects.filter(
        ejecutado_por__isnull=False,
        fecha_ejecucion__isnull=False,
        estado_suministro__estado_suministro__in=['EJECUTADO', 'DEVUELTO', 'ADMISIBLE'],
    ).exclude(ejecutado_por='')

    if fecha_inicio:
        suministros = suministros.filter(fecha_ejecucion__gte=fecha_inicio)
    if fecha_fin:
        suministros = suministros.filter(fecha_ejecucion__lte=fecha_fin)

    suministros = suministros.select_related('estado_suministro').order_by(
        'fecha_ejecucion', 'ejecutado_por'
    )

    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre',
    }

    matriz = defaultdict(
    lambda: defaultdict(
        lambda: {
            'monto': Decimal('0.00'),
            'ssts': set()
        }
    )
)
    ejecutores_set = set()
    totales_por_ejecutor = defaultdict(lambda: Decimal('0.00'))
    totales_por_fecha = defaultdict(lambda: Decimal('0.00'))
    total_general = Decimal('0.00')
    fechas_ordenadas = []
    fechas_formateadas = {}

    for s in suministros:
        fecha = s.fecha_ejecucion
        ejecutor = s.ejecutado_por.strip().title()
        monto = s.monto or Decimal('0.00')

        if fecha not in fechas_formateadas:
            fechas_formateadas[fecha] = f"{fecha.day:02d} de {meses[fecha.month]} de {fecha.year}"
            fechas_ordenadas.append(fecha)

        matriz[fecha][ejecutor]['monto'] += monto
        matriz[fecha][ejecutor]['ssts'].add(s.sst_id)
        ejecutores_set.add(ejecutor)
        totales_por_ejecutor[ejecutor] += monto
        totales_por_fecha[fecha] += monto
        total_general += monto

    ejecutores = sorted(ejecutores_set)
    fechas_ordenadas = sorted(set(fechas_ordenadas))

    reporte = [
    {
        'fecha': fechas_formateadas[fecha],

        'montos': [
            {
                'monto': matriz[fecha][ejecutor]['monto'],
                'cantidad': len(matriz[fecha][ejecutor]['ssts'])
            }
            for ejecutor in ejecutores
        ],

        'total': totales_por_fecha[fecha],
    }
    for fecha in fechas_ordenadas
]

    context = {
        'reporte': reporte,
        'ejecutores': ejecutores,
        'totales_por_ejecutor': [totales_por_ejecutor[e] for e in ejecutores],
        'total_general': total_general,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }
    return render(request, 'gestion/reporte_productividad.html', context)


# =============================================================================
# MAPA
# =============================================================================

def generar_colores_distintos(n):
    """Genera n colores visualmente distintos usando HSV."""
    colores = []
    for i in range(n):
        r, g, b = colorsys.hsv_to_rgb(i / n, 0.8, 0.9)
        colores.append('#{:02x}{:02x}{:02x}'.format(int(r * 255), int(g * 255), int(b * 255)))
    return colores


def mapa_suministros(request):
    qs = Suministro.objects.filter(
        estado_suministro__estado_suministro='ASIGNADO',
        latitud__isnull=False,
        longitud__isnull=False,
    ).exclude(latitud=0, longitud=0).select_related(
        'sst', 'estado_suministro'
    ).values(
        'suministro', 'direccion', 'latitud', 'longitud',
        'sst__sst', 'sst__id',
        'estado_suministro__estado_suministro', 'estado_suministro__color',
        'medidor', 'potencia', 'contacto', 'telefono',
    )

    ssts_unicos = list({s['sst__id'] for s in qs})
    mapa_colores = dict(zip(ssts_unicos, generar_colores_distintos(len(ssts_unicos))))

    suministros = []
    for s in qs:
        try:
            lat, lng = float(s['latitud']), float(s['longitud'])
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                suministros.append({
                    'suministro': s['suministro'],
                    'direccion': s['direccion'] or '',
                    'latitud': lat,
                    'longitud': lng,
                    'sst': s['sst__sst'] or '',
                    'sst_id': s['sst__id'],
                    'color': mapa_colores.get(s['sst__id'], '#3388ff'),
                    'estado': s['estado_suministro__estado_suministro'],
                    'estado_color': s['estado_suministro__color'] or '#3388ff',
                    'medidor': s['medidor'] or 'N/A',
                    'potencia': s['potencia'] or 'N/A',
                    'contacto': s['contacto'] or 'N/A',
                    'telefono': s['telefono'] or 'N/A',
                })
        except (ValueError, TypeError):
            continue

    ssts_info = {}
    for s in suministros:
        sst_id = s['sst_id']
        if sst_id not in ssts_info:
            ssts_info[sst_id] = {'nombre': s['sst'], 'color': s['color'], 'cantidad': 0}
        ssts_info[sst_id]['cantidad'] += 1

    return render(request, "gestion/mapa_suministros.html", {
        "suministros": json.dumps(suministros),
        "ssts_info": json.dumps(list(ssts_info.values())),
        "total": len(suministros),
    })


# =============================================================================
# LIQUIDACIÓN
# =============================================================================

def liquidacion_list(request):
    ssts = SST.objects.select_related(
        'estado_sst', 'estado_liquidacion', 'distrito'
    ).filter(estado_sst__estado='EJECUTADO').order_by('-created_at')

    ssts_data = []
    monto_total = Decimal('0.00')

    for sst in ssts:
        suministros = sst.suministros.all()
        ssts_data.append({
            'sst': sst,
            'ejecutados_por': ', '.join(
                filter(None, suministros.values_list('ejecutado_por', flat=True).distinct())
            ),
            'observaciones_contratista': ', '.join(
                filter(None, suministros.values_list('observacion_contratista', flat=True).distinct())
            ),
        })
        monto_total += sst.monto_proyectado or Decimal('0.00')

    estados_liquidacion = EstadoLiquidacion.objects.all()

    conteo_estados = [
        {
            'estado': estado.estado,
            'color': estado.color,
            'count': ssts.filter(estado_liquidacion=estado).count(),
            'monto': sum(
                s.monto_proyectado or Decimal('0.00')
                for s in ssts.filter(estado_liquidacion=estado)
            ),
        }
        for estado in estados_liquidacion
    ]

    return render(request, 'gestion/liquidacion_list.html', {
        'ssts_data': ssts_data,
        'estados_liquidacion': estados_liquidacion,
        'monto_total': monto_total,
        'conteo_estados': conteo_estados,
    })


@csrf_protect
@require_POST
def cambiar_estado_liquidacion(request, sst_id):
    sst = get_object_or_404(SST, id=sst_id)
    estado = get_object_or_404(EstadoLiquidacion, id=request.POST.get('estado_liquidacion_id'))
    sst.estado_liquidacion = estado
    sst.save()
    return JsonResponse({'success': True, 'estado': estado.estado, 'color': estado.color})


@csrf_protect
@require_POST
def actualizar_observacion_liquidacion(request, sst_id):
    sst = get_object_or_404(SST, id=sst_id)
    sst.observacion = request.POST.get('observacion', '')
    sst.save()
    return JsonResponse({'success': True})


@login_required
def descargar_excel_liquidacion(request):
    ssts = SST.objects.select_related(
        'estado_sst', 'estado_liquidacion', 'distrito'
    ).filter(estado_sst__estado='EJECUTADO').order_by('-created_at')

    wb = Workbook()
    ws = wb.active
    ws.title = "Liquidación"

    headers = [
        'SST', 'Distrito', 'Estado SST', 'Estado Liquidación',
        'Fecha Liq. Sistema', 'Monto Proyectado', 'Monto Real',
        'Ejecutado Por', 'Obs. Contratista', 'Observación',
    ]
    ws.append(headers)

    for sst in ssts:
        suministros = sst.suministros.all()
        ws.append([
            sst.sst,
            sst.distrito.nombre_distrito if sst.distrito else '',
            sst.estado_sst.estado if sst.estado_sst else '',
            sst.estado_liquidacion.estado if sst.estado_liquidacion else '',
            sst.fecha_liquidacion_sistema.strftime('%d/%m/%Y') if sst.fecha_liquidacion_sistema else '',
            float(sst.monto_proyectado or 0),
            float(sst.monto_real or 0),
            ', '.join(filter(None, suministros.values_list('ejecutado_por', flat=True).distinct())),
            ', '.join(filter(None, suministros.values_list('observacion_contratista', flat=True).distinct())),
            sst.observacion or '',
        ])

    filename = f"liquidacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# MANAGEMENT COMMAND (sincronizar montos)
# =============================================================================

class Command(BaseCommand):
    help = 'Sincroniza los montos de todas las SSTs con la suma de sus suministros'

    def handle(self, *args, **kwargs):
        ssts = SST.objects.all()
        total_actualizado = 0

        for sst in ssts:
            monto_anterior = sst.monto_proyectado
            monto_nuevo = sst.actualizar_monto_total()
            if monto_anterior != monto_nuevo:
                self.stdout.write(
                    self.style.SUCCESS(f'SST {sst.sst}: {monto_anterior} → {monto_nuevo}')
                )
                total_actualizado += 1

        self.stdout.write(
            self.style.SUCCESS(
                f'\n✅ {total_actualizado} SSTs actualizadas de {ssts.count()} totales'
            )
        )
        
        
        
 
 # views.py
import openpyxl
from django.http import HttpResponse
from .models import Suministro


def limpiar_codigo(valor):
    if not valor:
        return ""
    return str(valor)[:3].lstrip("0")


def descargar_para_modulo(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formato Modulo"

    headers = [
        "Item","Solicitud","Fecha Asignación","Actualizar a",
        "Alimentador","Sed","Suministro","Medidor Sel","Marca Sel",
        "Fase Sel","Potencia","Fecha Prim. Inst.","Area Solicitante",
        "Tipo Orden","Fecha Recepción","Nombre","Dirección","Distrito",
        "Sucursal","Fecha Instalación","Tipo Acometida Sel",
        "Tipo Tapa Sel","Tipo Caja Sel","Obs LDS","CR Responsable",
        "Contratista","Actividad 1","Actividad 2","Actividad 3","Actividad 4",
        "Cant FTE 1","Cant FTE 2","Cant FTE 3","Cant FTE 4",
        "Condición Mant.","Tip. Mot. Desestimado","Mot. Desestimado",
        "Fecha Ejecución","Ejecutado Por","Obs. Contra",
        "Fecha Rpta.","Usuario Rpta.","Medidor Enc.","Marca Enc.",
        "Fase Enc.","Lectura Enc.","Tipo Acometida Enc.",
        "Cod Des. Tip. Tapa Enc.","Cod Des. Tip. Caja Enc.",
        "Cant. Fotos Ant.","Ruta de Fotos Antes","Cant. Fotos Des.",
        "Ruta de Fotos Después","Ruta de Cuaderno Obra",
        "Ruta de Acta de Conformidad"
    ]

    ws.append(headers)

    registros = Suministro.objects.select_related(
        'estado_suministro', 'distrito', 'sst'
    ).all()

    for r in registros:

        estado_texto = r.estado_suministro.estado_suministro

        # 🔹 Transformación correcta
        if estado_texto == "EJECUTADO":
            estado_mod = "E=EJECUTADO"
        elif estado_texto == "DEVUELTO":
            estado_mod = "D=DEVUELTO"
        else:
            estado_mod = ""

        # Inicializar
        fte1 = fte2 = fte3 = fte4 = ""
        condicion = tip_mot = mot = ""
        fecha_ejec = ejecutado = obs_contra = ""

        # 🔥 EJECUTADO
        if estado_texto == "EJECUTADO":

            fte1 = limpiar_codigo(r.actividad1_prog)
            fte2 = limpiar_codigo(r.actividad2_prog)
            fte3 = limpiar_codigo(r.actividad3_prog)
            fte4 = limpiar_codigo(r.actividad4_prog)

            fecha_ejec = r.fecha_ejecucion
            ejecutado = r.ejecutado_por
            obs_contra = r.observacion_contratista

        # 🔥 DEVUELTO
        elif estado_texto == "DEVUELTO":

            condicion = "D=DEVUELTO"
            tip_mot = "020"
            mot = r.observacion_contratista

        fila = [
            r.item,
            r.sst.sst if r.sst else "",
            r.fecha_primer_envio,
            estado_mod,
            r.alim,
            r.sed,
            r.suministro,
            r.medidor,
            r.marca,
            r.fase,
            r.potencia,
            r.insta_med,
            r.area_solicitante,
            r.tipo_orden,
            r.fecha_asig,
            r.nombre,
            r.direccion,
            r.distrito.nombre_distrito if r.distrito else "",
            r.sucursal,
            r.insta_cnx,
            r.tipo_acometida,
            r.tipo_tapa,
            r.tipo_caja,
            r.Observacion_LDS,
            r.cr_emisor,
            r.contratista,

            r.actividad1_prog,
            r.actividad2_prog,
            r.actividad3_prog,
            r.actividad4_prog,

            fte1, fte2, fte3, fte4,

            condicion,
            tip_mot,
            mot,

            fecha_ejec,
            ejecutado,
            obs_contra,

            "", "", "", "", "", "", "", "", "",
            "", "", "", "", "", ""
        ]

        ws.append(fila)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=formato_modulo.xlsx'

    wb.save(response)
    return response